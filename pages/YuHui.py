import streamlit as st
import pandas as pd
import requests
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
import numpy as np
import datetime
import re
from openpyxl import load_workbook

st.set_page_config(page_title="YuHui Portfolio Dashboard", layout="wide")


def safe_secret(key: str, default=""):
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default


# =========================
# OneDrive Excel 同步（與上傳並存）
# =========================
def ensure_excel_from_onedrive(xlsx_path: Path) -> bool:
    url = safe_secret("ONEDRIVE_XLSX_URL", "")
    if not isinstance(url, str) or not url.strip():
        return False
    url = url.strip()

    def add_download_param(u: str) -> str:
        if "download=1" in u:
            return u
        return u + ("&" if "?" in u else "?") + "download=1"

    candidates = [url, add_download_param(url)]
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    last_err = None
    for u in candidates:
        try:
            r = requests.get(
                u,
                timeout=45,
                allow_redirects=True,
                headers={"User-Agent": "Mozilla/5.0"},
            )
            r.raise_for_status()
            content = r.content or b""
            if not content.startswith(b"PK"):
                last_err = RuntimeError(f"下載內容不是 Excel（前 20 bytes={content[:20]!r}）")
                continue
            xlsx_path.write_bytes(content)
            return True
        except Exception as e:
            last_err = e
            continue

    if last_err:
        st.warning(f"OneDrive 下載失敗，將使用既有檔案：{last_err}")
    return False


# =========================
# Google Drive / Google Sheets Excel 同步（與上傳並存）
# =========================
def _to_gdrive_xlsx_download_url(u: str) -> str | None:
    if not isinstance(u, str):
        return None
    u = u.strip()
    if not u:
        return None

    m = re.search(r"/spreadsheets/d/([^/]+)/", u)
    if m:
        sid = m.group(1)
        return f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"

    m = re.search(r"/file/d/([^/]+)/", u)
    if m:
        fid = m.group(1)
        return f"https://drive.google.com/uc?export=download&id={fid}"

    if "drive.google.com/uc" in u and "id=" in u:
        return u

    return None


def ensure_excel_from_gdrive(xlsx_path: Path) -> bool:
    raw = safe_secret("GOOGLE_SHEETS_URL", "") or safe_secret("GDRIVE_FILE_URL", "")
    if not isinstance(raw, str) or not raw.strip():
        return False

    url = _to_gdrive_xlsx_download_url(raw)
    if not url:
        st.warning("Google Drive 連結格式無法辨識，請確認是 Google Sheets 或 Drive 檔案分享連結。")
        return False

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        r = requests.get(url, timeout=45, allow_redirects=True, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        content = r.content or b""
        if not content.startswith(b"PK"):
            raise RuntimeError(f"下載內容不是 Excel（前 20 bytes={content[:20]!r}）")

        xlsx_path.write_bytes(content)
        return True
    except Exception as e:
        st.warning(f"Google Drive 下載失敗，將使用既有檔案：{e}")
        return False


def _touch_reload_flag(source: str):
    st.session_state["_reload_source"] = source


DATA_DIR = Path(__file__).resolve().parent.parent / "data"
XLSX_PATH = DATA_DIR / "yuhui_data.xlsx"


def to_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(
        s.astype(str)
         .str.replace(",", "", regex=False)
         .str.replace(" ", "", regex=False)
         .replace({"nan": None, "": None}),
        errors="coerce"
    ).fillna(0.0)


def _clean_text(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    return str(x).strip()


def extract_allocation_from_analysis_sheet(xlsx_path: Path, sheet_name: str = "YuHui"):
    """
    直接從 Excel 工作表讀『分析』區塊，避免 DataFrame 因合併儲存格/空白欄位而錯位。
    以 Excel 原始座標找：分析 / 分類 / 參考現值，因此能正確抓到台幣活儲、台股、台股 ETF、美股、美金儲蓄等列。
    """
    if not xlsx_path.exists():
        return None

    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception:
        return None

    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]

    def clean(v):
        if v is None:
            return ""
        return str(v).strip()

    anchor = None
    for row in ws.iter_rows():
        for cell in row:
            if clean(cell.value) == "分析":
                anchor = (cell.row, cell.column)
                break
        if anchor:
            break
    if not anchor:
        return None

    ar, ac = anchor

    def find_near(token, r0, r1, c0=None, c1=None):
        r0 = max(1, r0)
        r1 = min(ws.max_row, r1)
        c0 = 1 if c0 is None else max(1, c0)
        c1 = ws.max_column if c1 is None else min(ws.max_column, c1)
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                if clean(ws.cell(r, c).value) == token:
                    return r, c
        return None, None

    cat_r, cat_c = find_near("分類", ar - 3, ar + 12)
    if cat_r is None:
        return None

    val_r, val_c = find_near("參考現值", ar - 3, ar + 12)
    if val_r is None:
        val_r, val_c = find_near("成交金額", ar - 3, ar + 12)
    if val_r is None:
        return None

    items = []
    r = cat_r + 1
    while r <= ws.max_row:
        cat = clean(ws.cell(r, cat_c).value)
        if not cat:
            r += 1
            continue
        if cat == "總計":
            break

        raw = ws.cell(r, val_c).value
        val = pd.to_numeric(str(raw).replace(",", "").strip(), errors="coerce")
        if pd.isna(val):
            r += 1
            continue

        items.append({"分類": cat, "金額": float(val)})
        r += 1

    if not items:
        return None

    alloc = pd.DataFrame(items)
    alloc = alloc[alloc["金額"].fillna(0) != 0].copy()
    return alloc if not alloc.empty else None


def make_allocation_pie_from_analysis(xlsx_path: Path):
    alloc = extract_allocation_from_analysis_sheet(xlsx_path, sheet_name="YuHui")
    if alloc is None or alloc.empty:
        return None

    fig = px.pie(alloc, names="分類", values="金額", title="資產配置（依 Excel『分析』區塊）")
    fig.update_traces(textposition="inside", textinfo="percent+label")
    fig.update_layout(height=520)
    return fig



def make_holding_distribution_pie_by_market(family_df: pd.DataFrame, market: str):
    """
    持股分布圓餅圖：
    - 依『參考現值』計算
    - 台股：分類 = 台股 / 台股 ETF
    - 美股：分類 = 美股
    - 只計入參考現值 > 0 的目前持股
    - 同股票名稱自動合併
    """
    df = _filter_trade_like_rows(family_df).copy()

    if "分類" not in df.columns or "參考現值" not in df.columns:
        return None

    cat = df["分類"].astype(str).str.strip()
    if market == "台股":
        df = df[cat.isin(["台股", "台股 ETF"])]
    elif market == "美股":
        df = df[cat == "美股"]
    else:
        return None

    if df.empty:
        return None

    df["參考現值"] = to_num(df["參考現值"])
    df = df[df["參考現值"] > 0].copy()
    if df.empty:
        return None

    name_col = "股票名稱" if "股票名稱" in df.columns else ("股票" if "股票" in df.columns else None)
    code_col = "股票代號" if "股票代號" in df.columns else None
    if name_col is None:
        return None

    # 顯示名稱：有代號就用「代號 名稱」，避免同名或 ETF 名稱過長不好辨識
    if code_col:
        code = df[code_col].astype(str).str.strip().replace({"nan": ""})
        name = df[name_col].astype(str).str.strip().replace({"nan": ""})
        df["標的"] = np.where(code != "", code + " " + name, name)
    else:
        df["標的"] = df[name_col].astype(str).str.strip()

    agg = (
        df.groupby("標的", dropna=True)["參考現值"]
          .sum()
          .reset_index()
          .sort_values("參考現值", ascending=False)
    )

    if agg.empty:
        return None

    fig = px.pie(
        agg,
        names="標的",
        values="參考現值",
        title=f"持股分布（{market}，依參考現值）"
    )
    fig.update_traces(textposition="inside", textinfo="percent+label")
    fig.update_layout(height=520)
    return fig

def _filter_trade_like_rows(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    code_col = "股票代號" if "股票代號" in d.columns else None
    name_col = "股票名稱" if "股票名稱" in d.columns else ("股票" if "股票" in d.columns else None)

    if code_col:
        code = d[code_col].astype(str).str.strip()
        mask = code.notna() & (code != "") & (code.str.lower() != "nan")
        mask &= ~code.isin(["分類", "總計", "分析"])
        d = d[mask]
    elif name_col:
        name = d[name_col].astype(str).str.strip()
        mask = name.notna() & (name != "") & (name.str.lower() != "nan")
        mask &= ~name.isin(["分類", "總計", "分析"])
        d = d[mask]

    return d


@st.cache_data(show_spinner=False)
def load_data(xlsx_path: Path):
    xls = pd.ExcelFile(xlsx_path)
    family_df = pd.read_excel(xls, "YuHui")
    acct = pd.read_excel(xls, "YuHui-帳戶紀錄")
    return family_df, acct


def compute_kpi(family_df: pd.DataFrame):
    df = family_df.copy()
    try:
        df = _filter_trade_like_rows(df)
    except Exception:
        pass

    if "分類" in df.columns:
        cat = df["分類"].astype(str).str.strip()
        df = df[cat.notna() & (cat != "") & (cat.str.lower() != "nan")]

    invested_col = "成交金額"
    realized_col = "已實現損益"
    unrealized_col = "未實現損益"

    invested = to_num(df[invested_col]) if invested_col in df.columns else pd.Series([0.0])
    realized = to_num(df[realized_col]) if realized_col in df.columns else pd.Series([0.0])
    unrealized = to_num(df[unrealized_col]) if unrealized_col in df.columns else pd.Series([0.0])

    total_invested = float(invested.sum())
    total_realized = float(realized.sum())
    total_unrealized = float(unrealized.sum())
    total_pnl = total_realized + total_unrealized
    ret = (total_pnl / total_invested) if total_invested else 0.0

    return total_invested, total_realized, total_unrealized, total_pnl, ret



# ====== 進階績效：IRR / 資金使用率 / 10年預測 ======
def _first_existing_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    return next((c for c in candidates if c in df.columns), None)


def _xnpv(rate: float, cashflows: list[tuple[pd.Timestamp, float]]) -> float:
    if not cashflows:
        return 0.0
    d0 = min(d for d, _ in cashflows)
    total = 0.0
    for d, cf in cashflows:
        days = (d - d0).days
        total += cf / ((1.0 + rate) ** (days / 365.0))
    return total


def calc_xirr(cashflows: list[tuple[pd.Timestamp, float]]) -> float | None:
    """用日期現金流估算 XIRR。需同時有正、負現金流。"""
    cfs = [(pd.to_datetime(d), float(v)) for d, v in cashflows if pd.notna(d) and abs(float(v)) > 1e-9]
    if not cfs or not any(v > 0 for _, v in cfs) or not any(v < 0 for _, v in cfs):
        return None

    # 先用多組區間尋找符號反轉，再二分法；避免 numpy_financial 相依性
    grid = [-0.999, -0.9, -0.75, -0.5, -0.25, -0.1, 0.0, 0.05, 0.1, 0.2, 0.4, 0.7, 1.0, 1.5, 2.0, 3.0, 5.0, 10.0]
    vals = []
    for r in grid:
        try:
            v = _xnpv(r, cfs)
            if np.isfinite(v):
                vals.append((r, v))
        except Exception:
            pass

    for (lo, vlo), (hi, vhi) in zip(vals, vals[1:]):
        if vlo == 0:
            return lo
        if vlo * vhi > 0:
            continue
        for _ in range(100):
            mid = (lo + hi) / 2
            vmid = _xnpv(mid, cfs)
            if abs(vmid) < 1e-5:
                return mid
            if vlo * vmid <= 0:
                hi, vhi = mid, vmid
            else:
                lo, vlo = mid, vmid
        return (lo + hi) / 2
    return None


def build_investment_cashflows(richard: pd.DataFrame) -> list[tuple[pd.Timestamp, float]]:
    """
    以交易明細估算 IRR：
    - 買進日：成交金額視為現金流出
    - 賣出日：成交金額 + 已實現損益 + 除息 視為現金流入
    - 未賣出部位：今天以參考現值 + 除息 當作期末價值
    """
    df = _filter_trade_like_rows(richard).copy()
    if "分類" in df.columns:
        cat = df["分類"].astype(str).str.strip()
        df = df[cat.notna() & (cat != "") & (cat.str.lower() != "nan")]

    buy_date_col = _first_existing_col(df, ["買進日期", "日期"])
    sell_date_col = _first_existing_col(df, ["賣出日期", "出場日期"])
    amount_col = _first_existing_col(df, ["成交金額", "投入金額", "買進金額"])
    realized_col = _first_existing_col(df, ["已實現損益"])
    value_col = _first_existing_col(df, ["參考現值", "市值"])
    dividend_col = _first_existing_col(df, ["除息", "股息", "配息"])

    if buy_date_col is None or amount_col is None:
        return []

    today = pd.Timestamp.today().normalize()
    cashflows: list[tuple[pd.Timestamp, float]] = []
    for _, row in df.iterrows():
        buy_date = pd.to_datetime(row.get(buy_date_col), errors="coerce")
        amount = pd.to_numeric(str(row.get(amount_col, 0)).replace(",", ""), errors="coerce")
        if pd.notna(buy_date) and pd.notna(amount) and amount > 0:
            cashflows.append((buy_date, -float(amount)))

        div = pd.to_numeric(str(row.get(dividend_col, 0)).replace(",", ""), errors="coerce") if dividend_col else 0.0
        div = 0.0 if pd.isna(div) else float(div)

        sell_date = pd.to_datetime(row.get(sell_date_col), errors="coerce") if sell_date_col else pd.NaT
        if pd.notna(sell_date):
            realized = pd.to_numeric(str(row.get(realized_col, 0)).replace(",", ""), errors="coerce") if realized_col else 0.0
            realized = 0.0 if pd.isna(realized) else float(realized)
            cashflows.append((sell_date, float(amount) + realized + div))
        else:
            value = pd.to_numeric(str(row.get(value_col, 0)).replace(",", ""), errors="coerce") if value_col else 0.0
            value = 0.0 if pd.isna(value) else float(value)
            if value > 0 or div > 0:
                cashflows.append((today, value + div))
    return cashflows


def compute_advanced_metrics(richard: pd.DataFrame, acct: pd.DataFrame):
    total_invested, total_realized, total_unrealized, total_pnl, simple_ret = compute_kpi(richard)

    df = _filter_trade_like_rows(richard).copy()
    if "分類" in df.columns:
        cat = df["分類"].astype(str).str.strip()
        df = df[cat.notna() & (cat != "") & (cat.str.lower() != "nan")]

    value_col = _first_existing_col(df, ["參考現值", "市值"])
    current_invested_value = float(to_num(df[value_col]).sum()) if value_col else max(total_invested + total_unrealized, 0.0)

    cash_candidates = ["台幣現金水位", "台幣現金", "現金水位", "台幣結餘", "結餘"]
    date_col = "日期" if "日期" in acct.columns else acct.columns[0]
    acct2 = acct.copy()
    acct2[date_col] = pd.to_datetime(acct2[date_col], errors="coerce")
    acct2 = acct2.dropna(subset=[date_col]).sort_values(date_col)
    cash_col = _first_existing_col(acct2, cash_candidates)
    cash_balance = 0.0
    if cash_col and not acct2.empty:
        cash_balance = float(to_num(acct2[cash_col]).iloc[-1])

    total_assets = current_invested_value + cash_balance

    # ===== 資產績效：扣除後續新增本金，避免把「入金」誤當成「報酬」 =====
    # 優先使用帳戶紀錄的「台幣本金」作為累積投入本金；若沒有，才退回「投入金額 + 現金」。
    principal_candidates = ["台幣本金", "TWD本金", "本金(台幣)", "初始資金", "本金"]
    principal_col = _first_existing_col(acct2, principal_candidates)

    first_date = acct2[date_col].iloc[0] if not acct2.empty else pd.Timestamp.today().normalize()
    last_date = acct2[date_col].iloc[-1] if not acct2.empty else pd.Timestamp.today().normalize()

    initial_capital = 0.0
    total_contribution = 0.0
    if not acct2.empty:
        if principal_col:
            principal_series = to_num(acct2[principal_col])
            initial_capital = float(principal_series.iloc[0])
            total_contribution = float(principal_series.iloc[-1])
        elif cash_col:
            cash_series = to_num(acct2[cash_col])
            initial_capital = float(cash_series.iloc[0])
            total_contribution = total_invested + cash_balance

    # 若帳戶紀錄抓不到本金，保守退回「目前投入 + 現金」避免除以 0。
    if total_contribution <= 0:
        total_contribution = total_invested + cash_balance
    if initial_capital <= 0:
        initial_capital = total_contribution

    capital_usage = current_invested_value / total_assets if total_assets > 0 else 0.0

    # 資產報酬率：用目前總資產扣掉累積本金，只看真正增值。
    asset_gain = total_assets - total_contribution
    asset_return = (asset_gain / total_contribution) if total_contribution > 0 else 0.0

    # 年化資產報酬：用扣本金後的資產報酬做近似年化；若期間太短，避免失真就顯示 None。
    years = max((last_date - first_date).days / 365.0, 0.0)
    asset_cagr = None
    if years >= 0.25 and total_contribution > 0 and total_assets > 0:
        asset_cagr = (total_assets / total_contribution) ** (1 / years) - 1

    cfs = build_investment_cashflows(richard)
    irr = calc_xirr(cfs)

    # ===== 10年預測用：納入資金使用率後的有效年化報酬率 =====
    # 投資部位用 IRR；現金部位目前先視為 0% 報酬。若未來要納入定存/貨幣基金，可在這裡加入 cash_rate。
    investment_rate = irr if irr is not None and np.isfinite(irr) else 0.08
    cash_rate = 0.00
    effective_return_rate = investment_rate * capital_usage + cash_rate * (1 - capital_usage)

    # ===== 進階版 10年預測用：以「有效報酬 = IRR × 資金使用率」為來源，再做長期收斂 =====
    # raw effective_return_rate 反映目前短期績效；projection_base_rate 是拿來做長期預測的動態基準。
    # 預設用 0.5 收斂係數，避免短期績效直接外推 10 年造成過度樂觀。
    projection_convergence = 0.50
    projection_min_rate = 0.03
    projection_max_rate = 0.15
    projection_base_rate_raw = effective_return_rate * projection_convergence
    projection_base_rate = min(max(projection_base_rate_raw, projection_min_rate), projection_max_rate)

    # Alpha：投資 IRR 與整體資產年化報酬的落差，用來觀察資金使用效率是否拖累成果。
    alpha = None
    if irr is not None and asset_cagr is not None:
        alpha = irr - asset_cagr

    return {
        "irr": irr,
        "current_invested_value": current_invested_value,
        "cash_balance": cash_balance,
        "total_assets": total_assets,
        "initial_capital": initial_capital,
        "total_contribution": total_contribution,
        "asset_gain": asset_gain,
        "asset_return": asset_return,
        "asset_cagr": asset_cagr,
        "capital_usage": capital_usage,
        "effective_return_rate": effective_return_rate,
        "projection_convergence": projection_convergence,
        "projection_base_rate_raw": projection_base_rate_raw,
        "projection_base_rate": projection_base_rate,
        "projection_min_rate": projection_min_rate,
        "projection_max_rate": projection_max_rate,
        "alpha": alpha,
        "simple_ret": simple_ret,
    }

def make_10y_projection_chart(
    start_assets: float,
    effective_rate: float | None,
    annual_add: float = 0.0,
    convergence: float = 0.50,
    min_rate: float = 0.03,
    max_rate: float = 0.15,
):
    """
    進階版 10年資產預測：
    - 來源：有效年化報酬率 = IRR × 資金使用率 + 現金報酬 × 現金比例。
    - 長期預測基準：有效報酬 × 收斂係數，再套用上下限，避免短期 IRR 過度外推。
    - 三情境：保守 / 基準 / 樂觀，皆由動態基準推導，不再是固定寫死。
    """
    if start_assets <= 0:
        return None
    if effective_rate is None or not np.isfinite(effective_rate):
        effective_rate = 0.08

    dynamic_base_raw = effective_rate * convergence
    base_rate = min(max(dynamic_base_raw, min_rate), max_rate)
    conservative_rate = min(max(base_rate * 0.70, min_rate), max_rate)
    optimistic_rate = min(max(base_rate * 1.30, min_rate), max_rate)

    scenarios = {
        f"保守 {conservative_rate*100:.1f}%": conservative_rate,
        f"基準 {base_rate*100:.1f}%（有效報酬×{convergence:.2f}）": base_rate,
        f"樂觀 {optimistic_rate*100:.1f}%": optimistic_rate,
    }
    rows = []
    for name, r in scenarios.items():
        value = start_assets
        for y in range(0, 11):
            if y == 0:
                value = start_assets
            else:
                value = value * (1 + r) + annual_add
            rows.append({"年度": y, "情境": name, "預測資產": value})
    dfp = pd.DataFrame(rows)
    fig = px.line(
        dfp,
        x="年度",
        y="預測資產",
        color="情境",
        markers=True,
        title="10年資產預測（有效報酬動態基準 + 長期收斂）",
    )
    fig.update_layout(height=520, yaxis_title="資產金額", legend_title_text="")
    fig.update_yaxes(tickformat=",")
    return fig


def make_rank_chart_by_market(family_df: pd.DataFrame, market: str, top_n: int = 10):
    realized_col = "已實現損益"
    unrealized_col = "未實現損益"
    cat_col = "分類"

    df = _filter_trade_like_rows(family_df)

    name_col = "股票名稱" if "股票名稱" in df.columns else ("股票" if "股票" in df.columns else None)
    if name_col is None:
        return None

    if cat_col in df.columns:
        cat = df[cat_col].astype(str).str.strip()
        if market == "美股":
            df = df[cat == "美股"]
        else:
            df = df[cat.isin(["台股", "台股 ETF"])]
    else:
        return None

    if df.empty:
        return None

    df["已實現損益"] = to_num(df[realized_col]) if realized_col in df.columns else 0.0
    df["未實現損益"] = to_num(df[unrealized_col]) if unrealized_col in df.columns else 0.0
    df["總損益"] = df["已實現損益"] + df["未實現損益"]

    agg = (
        df.groupby(name_col, dropna=True)["總損益"]
          .sum()
          .sort_values(ascending=False)
          .head(top_n)
          .reset_index()
          .rename(columns={name_col: "股票", "總損益": "總損益"})
    )

    if agg.empty:
        return None

    bar_colors = np.where(agg["總損益"] >= 0, "#1f77b4", "#d62728")
    bar_text = agg["總損益"].map(lambda v: f"{v:,.0f}")

    fig = go.Figure()
    fig.add_bar(
        x=agg["總損益"],
        y=agg["股票"],
        orientation="h",
        name="總損益",
        marker_color=bar_colors,
        text=bar_text,
        textposition="outside",
    )

    fig.update_layout(
        title=f"{market} 股票別總損益 Top {top_n}",
        height=520,
        margin=dict(t=70),
        showlegend=False,
    )
    fig.update_xaxes(title="總損益", zeroline=True, zerolinewidth=1, zerolinecolor="gray")
    fig.update_yaxes(title="股票", categoryorder="total ascending")
    return fig


def make_timeseries(acct: pd.DataFrame):
    date_col = "日期" if "日期" in acct.columns else acct.columns[0]
    df0 = acct.copy()
    df0[date_col] = pd.to_datetime(df0[date_col], errors="coerce")
    df0 = df0.dropna(subset=[date_col]).sort_values(date_col)

    principal_candidates = ["台幣本金", "TWD本金", "本金(台幣)"]
    cash_candidates = ["台幣現金水位", "台幣現金", "現金水位", "台幣結餘", "結餘"]

    principal_col = next((c for c in principal_candidates if c in df0.columns), None)
    cash_col = next((c for c in cash_candidates if c in df0.columns), None)

    if principal_col is None and cash_col is None:
        return None

    parts = []
    if cash_col is not None:
        tmp = df0[[date_col, cash_col]].copy()
        tmp["值"] = to_num(tmp[cash_col])
        tmp["項目"] = "台幣現金水位"
        parts.append(tmp[[date_col, "值", "項目"]])

    if principal_col is not None:
        tmp = df0[[date_col, principal_col]].copy()
        tmp["值"] = to_num(tmp[principal_col])
        tmp["項目"] = "台幣本金"
        parts.append(tmp[[date_col, "值", "項目"]])

    df = pd.concat(parts, ignore_index=True)

    if df["項目"].nunique() == 1:
        only = df["項目"].iloc[0]
        fig = px.line(df, x=date_col, y="值", title=f"台幣現金水位圖（來源：帳戶紀錄 / {only}）")
        fig.update_layout(height=450, yaxis_title=only, legend_title_text="")
        return fig

    fig = px.line(df, x=date_col, y="值", color="項目", title="台幣現金水位圖（台幣現金水位 vs 台幣本金）")
    fig.update_layout(height=450, legend_title_text="")
    fig.update_yaxes(title_text="金額")
    return fig


def make_yearly_return_combo(family_df: pd.DataFrame, mode: str = "已實現", attrib: str = "A"):
    realized_col = "已實現損益"
    unrealized_col = "未實現損益"

    buy_date_col = "買進日期" if "買進日期" in family_df.columns else None
    sell_date_col = "賣出日期" if "賣出日期" in family_df.columns else None

    if realized_col not in family_df.columns:
        return None

    df = _filter_trade_like_rows(family_df).copy()

    current_year = datetime.date.today().year
    min_year = 2000
    max_year = current_year

    def _clean_year_series(s: pd.Series) -> pd.Series:
        y = pd.to_numeric(s, errors="coerce")
        return y.where((y >= min_year) & (y <= max_year))

    if buy_date_col is not None:
        df[buy_date_col] = pd.to_datetime(df[buy_date_col], errors="coerce")
    if sell_date_col is not None:
        df[sell_date_col] = pd.to_datetime(df[sell_date_col], errors="coerce")

    sold = df.copy()
    if sell_date_col is not None:
        sold = sold[sold[sell_date_col].notna()].copy()

    yearly_realized = None

    if mode == "已實現":
        if attrib == "A":
            if sell_date_col is None:
                return None
            sold["年度"] = _clean_year_series(sold[sell_date_col].dt.year)
            sold = sold[sold["年度"].notna()].copy()
            sold["年度收益"] = to_num(sold[realized_col])
            yearly_realized = sold.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")

        elif attrib == "B":
            if buy_date_col is None:
                return None
            sold = sold[sold[buy_date_col].notna()].copy()
            sold["年度"] = _clean_year_series(sold[buy_date_col].dt.year)
            sold = sold[sold["年度"].notna()].copy()
            sold["年度收益"] = to_num(sold[realized_col])
            yearly_realized = sold.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")

        elif attrib == "C":
            if buy_date_col is None or sell_date_col is None:
                return None

            d = sold[sold[buy_date_col].notna() & sold[sell_date_col].notna()].copy()
            if d.empty:
                return None

            pnl = to_num(d[realized_col]).fillna(0.0).to_numpy()
            rows = []

            for i, r in enumerate(d.itertuples(index=False)):
                b = getattr(r, buy_date_col)
                s = getattr(r, sell_date_col)
                if pd.isna(b) or pd.isna(s):
                    continue

                b = pd.Timestamp(b).normalize()
                s = pd.Timestamp(s).normalize()

                if b.year < min_year:
                    b = pd.Timestamp(f"{min_year}-01-01")
                if s.year > max_year:
                    s = pd.Timestamp(f"{max_year}-12-31")

                if s.year < min_year or b.year > max_year:
                    continue

                if s < b:
                    y = s.year
                    if min_year <= y <= max_year:
                        rows.append((y, float(pnl[i])))
                    continue

                total_days = max((s - b).days + 1, 1)
                start_year = max(b.year, min_year)
                end_year = min(s.year, max_year)

                for y in range(start_year, end_year + 1):
                    seg_start = max(b, pd.Timestamp(f"{y}-01-01"))
                    seg_end = min(s, pd.Timestamp(f"{y}-12-31"))
                    seg_days = (seg_end - seg_start).days + 1
                    if seg_days <= 0:
                        continue
                    rows.append((y, float(pnl[i]) * (seg_days / total_days)))

            if not rows:
                return None

            tmp = pd.DataFrame(rows, columns=["年度", "年度收益"])
            tmp["年度"] = _clean_year_series(tmp["年度"])
            tmp = tmp[tmp["年度"].notna()].copy()
            yearly_realized = tmp.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")

        else:
            return None

        yearly = yearly_realized

    else:
        if sell_date_col is None:
            return None

        sold2 = df[df[sell_date_col].notna()].copy()
        sold2["年度"] = _clean_year_series(sold2[sell_date_col].dt.year)
        sold2 = sold2[sold2["年度"].notna()].copy()
        sold2["年度收益"] = to_num(sold2[realized_col])
        yearly_realized = sold2.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")

        open_pos = df[df[sell_date_col].isna()].copy()
        if unrealized_col in open_pos.columns and not open_pos.empty:
            open_pos["年度"] = current_year
            open_pos["年度收益"] = to_num(open_pos[unrealized_col])
            yearly_unrealized = open_pos.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")
            yearly = pd.concat([yearly_realized, yearly_unrealized], ignore_index=True)
            yearly = yearly.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")
        else:
            yearly = yearly_realized

    if yearly is None or yearly.empty:
        return None

    yearly["年度"] = _clean_year_series(yearly["年度"])
    yearly = yearly[yearly["年度"].notna()].copy()
    if yearly.empty:
        return None

    yearly = yearly.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")
    yearly["年度"] = yearly["年度"].astype(int)
    yearly["累積收益"] = yearly["年度收益"].cumsum()
    yearly["累積標籤"] = yearly["累積收益"].map(lambda v: f"{v:,.0f}")
    yearly["年度標籤"] = yearly["年度收益"].map(lambda v: f"{v:,.0f}")

    fig = go.Figure()
    bar_colors = np.where(yearly["年度收益"] >= 0, "#1f77b4", "#d62728")
    bar_text_pos = ["outside" if v >= 0 else "inside" for v in yearly["年度收益"]]

    fig.add_bar(
        x=yearly["年度"].astype(str),
        y=yearly["年度收益"],
        name="年度收益",
        marker_color=bar_colors,
        text=yearly["年度標籤"],
        textposition=bar_text_pos,
        yaxis="y",
    )

    fig.add_trace(go.Scatter(
        x=yearly["年度"].astype(str),
        y=yearly["累積收益"],
        name="累積收益",
        mode="lines+markers+text",
        text=yearly["累積標籤"],
        textposition="top center",
        yaxis="y2"
    ))

    left_max = float(max(yearly["年度收益"].max(), 0))
    left_min = float(min(yearly["年度收益"].min(), 0))
    left_pad = max((left_max - left_min) * 0.15, 1000)

    right_max = float(yearly["累積收益"].max())
    right_min = float(yearly["累積收益"].min())
    right_pad = max((right_max - right_min) * 0.15, 1000)

    title_suffix = mode if mode != "已實現" else f"{mode}（{attrib}）"
    fig.update_layout(
        title=f"投資收益（年度 vs 累積）— {title_suffix}",
        xaxis=dict(title="年度", type="category"),
        yaxis=dict(title="年度收益", range=[left_min - left_pad, left_max + left_pad]),
        yaxis2=dict(
            title="累積收益",
            overlaying="y",
            side="right",
            showgrid=False,
            range=[right_min - right_pad, right_max + right_pad],
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
        height=520,
        margin=dict(t=80)
    )
    return fig


view_mode = st.sidebar.radio("顯示內容", ["圖表", "交易明細"], index=0)

source = st.session_state.pop("_reload_source", None)
need_fetch = (not XLSX_PATH.exists()) or (source in ("gdrive", "onedrive"))
if need_fetch:
    fetched = False
    if source == "gdrive":
        fetched = ensure_excel_from_gdrive(XLSX_PATH)
        if (not fetched) and safe_secret("ONEDRIVE_XLSX_URL"):
            fetched = ensure_excel_from_onedrive(XLSX_PATH)
    elif source == "onedrive":
        fetched = ensure_excel_from_onedrive(XLSX_PATH)
        if (not fetched) and (safe_secret("GOOGLE_SHEETS_URL") or safe_secret("GDRIVE_FILE_URL")):
            fetched = ensure_excel_from_gdrive(XLSX_PATH)
    else:
        if safe_secret("GOOGLE_SHEETS_URL") or safe_secret("GDRIVE_FILE_URL"):
            fetched = ensure_excel_from_gdrive(XLSX_PATH)
        if (not fetched) and safe_secret("ONEDRIVE_XLSX_URL"):
            fetched = ensure_excel_from_onedrive(XLSX_PATH)

if not XLSX_PATH.exists():
    st.error("找不到 data/yuhui_data.xlsx。")
    st.stop()

family_df, acct = load_data(XLSX_PATH)

st.title("YuHui 的投資儀表板")
if XLSX_PATH.exists():
    st.caption(f"資料最後更新時間：{pd.to_datetime(XLSX_PATH.stat().st_mtime, unit='s')}")

total_invested, total_realized, total_unrealized, total_pnl, ret = compute_kpi(family_df)
adv = compute_advanced_metrics(family_df, acct)

# ====== KPI 說明小工具 ======
# Streamlit 的 metric 支援 help 參數；滑鼠移到指標旁的 ? 可看到公式與用途。
def kpi_metric(col, label: str, value: str, help_text: str):
    try:
        col.metric(label, value, help=help_text)
    except TypeError:
        # 舊版 Streamlit 若不支援 help，仍可正常顯示數值。
        col.metric(label, value)

alpha_text = "資料不足" if adv["alpha"] is None else f"{adv['alpha']*100:,.2f}%"
irr_text = "資料不足" if adv["irr"] is None else f"{adv['irr']*100:,.2f}%"
asset_cagr_text = "資料不足" if adv["asset_cagr"] is None else f"{adv['asset_cagr']*100:,.2f}%"

c1, c2, c3, c4, c5 = st.columns(5)
kpi_metric(c1, "投入金額", f"{total_invested:,.0f}", "交易明細中的累積成交金額。公式：Σ 成交金額。用途：作為投資部位的投入本金基準。")
kpi_metric(c2, "已實現損益", f"{total_realized:,.0f}", "已賣出部位已經落袋的損益。公式：Σ 已實現損益。")
kpi_metric(c3, "未實現損益", f"{total_unrealized:,.0f}", "尚未賣出部位依目前參考現值估算的浮動損益。公式：Σ 未實現損益。")
kpi_metric(c4, "總損益", f"{total_pnl:,.0f}", "整體投資損益。公式：已實現損益 + 未實現損益。")
kpi_metric(c5, "報酬率", f"{ret*100:,.2f}%", "投資部位的簡單報酬率，不含時間因素。公式：總損益 ÷ 投入金額。")

c6, c7, c8, c9, c10 = st.columns(5)
kpi_metric(c6, "年化報酬率 IRR", irr_text, "投資現金流的年化報酬率，考慮買進日、賣出日與未賣出部位參考現值。用途：衡量投資操作能力。")
kpi_metric(c7, "總資產", f"{adv['total_assets']:,.0f}", "目前總資產。公式：投資部位現值 + 現金水位。")
kpi_metric(c8, "年化資產報酬", asset_cagr_text, "整體財富的年化成長速度，已扣除累積本金影響。公式概念：(總資產 ÷ 累積本金)^(1/年數) - 1。用途：衡量真正變有錢的速度。")
kpi_metric(c9, "資金使用率", f"{adv['capital_usage']*100:,.2f}%", "目前有多少資產實際投入市場。公式：投資部位現值 ÷ 總資產。")
kpi_metric(c10, "現金水位", f"{adv['cash_balance']:,.0f}", "帳戶紀錄中最新一筆台幣現金水位。用途：評估可加碼資金與防守能力。")

c11, c12, c13, c14, c15 = st.columns(5)
kpi_metric(c11, "累積本金", f"{adv['total_contribution']:,.0f}", "目前用來衡量資產成長的本金基準。通常取帳戶紀錄中的資金累積值，避免把後續入金誤算成投資報酬。")
kpi_metric(c12, "資產增值", f"{adv['asset_gain']:,.0f}", "扣除累積本金後真正增加的資產。公式：總資產 - 累積本金。")
kpi_metric(c13, "資產報酬率", f"{adv['asset_return']*100:,.2f}%", "整體資產相對累積本金的累積報酬，不年化。公式：資產增值 ÷ 累積本金。")
kpi_metric(c14, "有效年化報酬率", f"{adv['effective_return_rate']*100:,.2f}%", "把資金使用率納入後的投資效率。公式：IRR × 資金使用率。用途：避免只看投資部位 IRR 而忽略閒置現金。")
kpi_metric(c15, "IRR-資產年化差", alpha_text, "投資部位年化報酬與整體資產年化報酬的差距。公式：IRR - 年化資產報酬。差距越大，通常代表現金閒置或資金配置效率仍有改善空間。")

st.caption("註：滑鼠移到各項指標旁的說明圖示可查看公式與用途；10年預測來源為『有效年化報酬率 = IRR × 資金使用率』，再乘以收斂係數並套用上下限，避免短期 IRR 直接外推。")
st.divider()


def render_trade_details(family_df: pd.DataFrame):
    st.subheader("交易明細")

    if "分類" not in family_df.columns:
        st.warning("找不到『分類』欄位，無法依台股/美股切換。")
        st.dataframe(family_df, use_container_width=True)
        return

    market = st.radio("明細篩選", ["台股（含台股 ETF）", "美股", "全部"], horizontal=True)

    df = family_df.copy()
    try:
        df = _filter_trade_like_rows(df)
    except Exception:
        pass

    cat = df["分類"].astype(str).str.strip()
    if market.startswith("台股"):
        df = df[cat.isin(["台股", "台股 ETF"])]
    elif market == "美股":
        df = df[cat == "美股"]
    else:
        df = df[cat.notna() & (cat != "") & (cat.str.lower() != "nan")]

    preferred_cols = [
        "買進日期", "賣出日期", "股票代號", "股票名稱", "分類",
        "股數", "買進價", "賣出價",
        "成交金額", "手續費", "交易稅", "除息",
        "已實現損益", "未實現損益", "參考現值",
        "買進原因", "賣出原因", "備註"
    ]
    cols = [c for c in preferred_cols if c in df.columns]
    df_view = df[cols] if cols else df

    for dc in ["買進日期", "賣出日期"]:
        if dc in df_view.columns:
            df_view[dc] = pd.to_datetime(df_view[dc], errors="coerce")

    st.dataframe(df_view, use_container_width=True, height=560)
    csv = df_view.to_csv(index=False, encoding="utf-8-sig")
    st.download_button("下載明細 CSV", data=csv, file_name="trades.csv", mime="text/csv")


if view_mode == "圖表":
    mode = st.radio("年度收益模式", ["已實現", "含未實現"], horizontal=True)
    attrib = st.radio("年度歸類方式（已實現用）", ["A 賣出年度（實現制）", "B 買進年度（決策歸因）", "C 跨年度攤提（天數分攤）"], horizontal=True)
    attrib_key = attrib.split()[0]

    yearly_fig = make_yearly_return_combo(family_df, mode=mode, attrib=attrib_key)
    if yearly_fig is not None:
        st.plotly_chart(yearly_fig, use_container_width=True)
    else:
        st.info("無法產生『投資收益（年度 vs 累積）』圖表（請確認 Excel 有『賣出日期 / 已實現損益』）。")

    st.subheader("10年資產預測")
    default_monthly = 0
    monthly_add = st.number_input("每月新增投入金額（可自行調整）", min_value=0, value=default_monthly, step=1000)

    with st.expander("進階預測參數（可調整預測基準）", expanded=False):
        projection_convergence = st.slider(
            "有效報酬收斂係數",
            min_value=0.20,
            max_value=1.00,
            value=float(adv["projection_convergence"]),
            step=0.05,
            help="1.00 代表完全採用目前有效報酬；0.50 代表只採用一半，較適合長期預測。",
        )
        max_projection_rate_pct = st.slider(
            "預測報酬率上限",
            min_value=8.0,
            max_value=25.0,
            value=float(adv["projection_max_rate"] * 100),
            step=0.5,
            help="避免短期高 IRR 讓 10 年預測過度膨脹。",
        )
        min_projection_rate_pct = st.slider(
            "預測報酬率下限",
            min_value=0.0,
            max_value=8.0,
            value=float(adv["projection_min_rate"] * 100),
            step=0.5,
        )

        preview_base_raw = adv["effective_return_rate"] * projection_convergence
        preview_base = min(max(preview_base_raw, min_projection_rate_pct / 100), max_projection_rate_pct / 100)
        st.write(
            f"目前有效年化報酬率：{adv['effective_return_rate']*100:.2f}% → "
            f"收斂後原始基準：{preview_base_raw*100:.2f}% → "
            f"套用上下限後基準：{preview_base*100:.2f}%"
        )

    proj_fig = make_10y_projection_chart(
        adv["total_assets"],
        adv["effective_return_rate"],
        annual_add=monthly_add * 12,
        convergence=projection_convergence,
        min_rate=min_projection_rate_pct / 100,
        max_rate=max_projection_rate_pct / 100,
    )
    if proj_fig is not None:
        st.plotly_chart(proj_fig, use_container_width=True)
    else:
        st.info("總資產資料不足，無法產生 10 年資產預測。")

    pie = make_allocation_pie_from_analysis(XLSX_PATH)
    if pie is not None:
        st.plotly_chart(pie, use_container_width=True)
    else:
        st.warning("找不到 Excel 內『分析』區塊（含『分類』與『參考現值』）或該區塊資料為空。")

    # ====== 持股分布（台股 / 美股分開）======
    st.subheader("持股分布（依參考現值）")
    hold_col1, hold_col2 = st.columns(2)

    with hold_col1:
        tw_hold_fig = make_holding_distribution_pie_by_market(family_df, "台股")
        if tw_hold_fig is not None:
            st.plotly_chart(tw_hold_fig, use_container_width=True)
        else:
            st.info("沒有可用的台股持股分布資料。")

    with hold_col2:
        us_hold_fig = make_holding_distribution_pie_by_market(family_df, "美股")
        if us_hold_fig is not None:
            st.plotly_chart(us_hold_fig, use_container_width=True)
        else:
            st.info("沒有可用的美股持股分布資料。")

    ts = make_timeseries(acct)
    if ts is not None:
        st.plotly_chart(ts, use_container_width=True)
    else:
        st.warning("帳戶紀錄缺少台幣現金相關欄位（台幣現金/台幣本金/結餘），無法畫台幣現金水位圖。")

    top_market = st.radio("Top10 類型", ["台股（含台股 ETF）", "美股"], horizontal=True)
    if top_market.startswith("台股"):
        top_fig = make_rank_chart_by_market(family_df, market="台股", top_n=10)
        if top_fig is not None:
            st.plotly_chart(top_fig, use_container_width=True)
        else:
            st.info("沒有找到可用的台股資料（Top 10）。")
    else:
        top_fig = make_rank_chart_by_market(family_df, market="美股", top_n=10)
        if top_fig is not None:
            st.plotly_chart(top_fig, use_container_width=True)
        else:
            st.info("沒有找到可用的美股資料（Top 10）。")
else:
    render_trade_details(family_df)
